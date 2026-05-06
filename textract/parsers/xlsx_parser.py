import openpyxl

from .utils import BaseParser


class Parser(BaseParser):
    """Extract text from Excel files (.xls/.xlsx)."""

    def extract(self, filename, **kwargs):
        workbook = openpyxl.load_workbook(filename, data_only=True)
        output = "\n"
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            for row in worksheet.iter_rows(values_only=True):
                
                non_empty_values = []
                for value in row:
                    if value:
                        if isinstance(value, bool):
                            value = "1"
                        elif isinstance(value, (int, float)):
                            # Convert to float to preserve decimal format (e.g., 83 -> 83.0)
                            value = str(float(value))
                            non_empty_values.append(value)
                        else:
                            non_empty_values.append(str(value))
                        non_empty_values.append(str(value))
                if non_empty_values:
                    output += " ".join(non_empty_values) + "\n"
        return output
